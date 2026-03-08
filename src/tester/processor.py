"""
Tester Tool - 数据处理模块
"""

import re
from pathlib import Path
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def is_merged_cell(ws, row, col):
    """检查单元格是否是合并单元格"""
    coord = ws.cell(row, col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if coord in merged_range:
            return True
    return False


class DataProcessor:
    """试验数据处理器"""
    
    def __init__(self):
        self.csv_data = {}  # {file_num: {channel: [(time, value), ...]}}
        self.warnings = []  # 存储警告信息
    
    def process(self, csv_dir: str = None, csv_files: list = None, excel_file: str = None, ambient_cols: list = None,
                file_index_col: int = 4, channel_index_col: int = 5,
                time_interval: int = 60, temp_threshold: float = 2.0, 
                temp_threshold_step: float = 0.5, log_callback=None) -> dict:
        """
        处理数据
        
        Args:
            csv_dir: CSV 根目录（包含1/2/3子目录）
            csv_files: CSV 文件列表（兼容旧版）
            excel_file: Excel 模板文件路径
            ambient_cols: 用户选择的环境温度行信息 [{'row': int, 'd_value': str, 'e_value': str}, ...]
            file_index_col: 文件索引列号（默认4=D列）
            channel_index_col: 通道索引列号（默认5=E列）
            time_interval: 稳定时间间隔（分钟）
            temp_threshold: 初始温差阈值（°C）
            temp_threshold_step: 温差阈值调整步长（°C）
            
        Returns:
            dict: {'message': str, 'output_path': str, 'warnings': list}
        """
        self.warnings = []  # 重置警告
        self.logs = []  # 重置日志
        
        # 优先使用 csv_dir，兼容 csv_files
        if csv_dir:
            csv_files = self.parse_csv_directory(csv_dir)
        
        if not csv_files:
            raise ValueError("目录检查阶段: 未找到 CSV 文件，请检查目录结构")
        
        # 解析 CSV 文件
        csv_start, csv_end = None, None
        
        if csv_dir:
            # 新版：从目录结构解析，文件索引由 Excel D 列决定
            # 先建立子目录名 -> CSV 文件路径的映射
            subdir_to_csv = {}
            root_path = Path(csv_dir)
            for subdir in root_path.iterdir():
                if subdir.is_dir() and subdir.name.isdigit():
                    csv_list = list(subdir.glob("*.csv"))
                    if csv_list:
                        subdir_to_csv[subdir.name] = str(csv_list[0])
            
            # 解析所有 CSV 文件，获取时间范围
            for subdir_name, csv_path in subdir_to_csv.items():
                file_num = int(subdir_name)
                data, start, end = self.parse_csv(csv_path)
                self.csv_data[file_num] = data
                
                if csv_start is None or (start and start < csv_start):
                    csv_start = start
                if csv_end is None or (end and end > csv_end):
                    csv_end = end
        else:
            # 兼容旧版：从文件列表直接解析
            subdir_to_csv = None
            for csv_path in csv_files:
                file_num = int(Path(csv_path).stem)
                data, start, end = self.parse_csv(csv_path)
                self.csv_data[file_num] = data
                
                if csv_start is None or (start and start < csv_start):
                    csv_start = start
                if csv_end is None or (end and end > csv_end):
                    csv_end = end
        
        if not csv_start or not csv_end:
            raise ValueError("CSV解析阶段: 无法解析 CSV 文件的时间数据")
        
        # 加载 Excel
        wb = load_workbook(excel_file)
        ws = wb.active
        
        # 更新表头
        ws.cell(1, 1).value = f"Beginning time: {csv_start.strftime('%H:%M')}              Ending time: {csv_end.strftime('%H:%M')}"
        
        duration = csv_end - csv_start
        hours = int(duration.total_seconds() // 3600)
        minutes = int((duration.total_seconds() % 3600) // 60)
        ws.cell(1, 4).value = f"Test duration:{hours}h{minutes}min"
        ws.cell(2, 1).value = f"Test date: {csv_start.strftime('%Y-%m-%d')}"
        
        # 获取所有时间点
        all_times = set()
        for fn in self.csv_data:
            for ch in self.csv_data[fn]:
                for t, _ in self.csv_data[fn][ch]:
                    all_times.add(t)
        sorted_times = sorted(all_times)
        
        # 获取环境温度信息（直接使用GUI传递的数据）
        ambient_info = []
        for col_info in ambient_cols:
            # 直接使用传递过来的 d_value 和 e_value
            d_val = col_info.get('d_value')
            e_val = col_info.get('e_value')
            
            if d_val and e_val:
                try:
                    file_num = int(float(str(d_val)))
                    channel = int(float(str(e_val)))
                    ambient_info.append((file_num, channel))
                except:
                    pass
        
        # 收集所有数据行（使用用户选择的索引列）
        data_rows = []
        skip_channels = set()
        
        # 从Excel读取所有数据行（根据索引列）
        for row in range(4, ws.max_row + 1):
            d_val = ws.cell(row, file_index_col).value  # 用户选择的文件索引列
            e_val = ws.cell(row, channel_index_col).value  # 用户选择的通道索引列
            i_val = ws.cell(row, 9).value  # I列 - ref标记
            limit_val = ws.cell(row, 10).value  # J列 - Limit
            b_val = ws.cell(row, 2).value  # B列 - 名称
            
            if d_val and e_val:
                try:
                    # 这里 d_val 现在是子目录名（如 "1", "2", "3"）
                    file_num = int(float(str(d_val)))
                    channel = int(float(str(e_val)))
                    
                    # 检查是否跳过（ref标记）
                    i_str = str(i_val).strip().lower() if i_val else ''
                    if 'ref' in i_str or i_str in ['--', '—', '-']:
                        skip_channels.add((file_num, channel))
                    
                    # 检查数据范围
                    if file_num in self.csv_data and channel in self.csv_data[file_num]:
                        for _, val in self.csv_data[file_num][channel]:
                            if val < -100 or val > 500:
                                skip_channels.add((file_num, channel))
                                break
                    
                    # 存储limit值
                    try:
                        limit_num = float(str(limit_val)) if limit_val else None
                    except:
                        limit_num = None
                    
                    data_rows.append({
                        'row': row,
                        'file': file_num,
                        'channel': channel,
                        'limit': limit_num,
                        'b_value': str(b_val).strip() if b_val else ''
                    })
                except:
                    pass
        
        # 查找稳定时刻（支持逐步增加阈值）
        check_channels = [
            (dr['file'], dr['channel'])
            for dr in data_rows
            if (dr['file'], dr['channel']) not in skip_channels
        ]
        
        # 使用逐步增加阈值的方式查找稳定时间
        initial_threshold = temp_threshold
        actual_threshold = temp_threshold
        stable_start = None
        stable_end = None
        
        # 最大尝试次数，防止无限循环
        max_attempts = 20
        
        def emit_log(msg):
            """输出日志（实时或批量）"""
            self.logs.append(msg)
            if log_callback:
                log_callback(msg)
        
        while actual_threshold <= initial_threshold + temp_threshold_step * max_attempts:
            if actual_threshold > initial_threshold:
                # 记录尝试日志
                next_threshold = actual_threshold + temp_threshold_step
                emit_log(f"温差阈值 {round(actual_threshold, 2)}K 不满足条件，尝试 {round(next_threshold, 2)}K...")
            
            stable_start, stable_end = self.find_stable_time(
                sorted_times, ambient_info, check_channels, time_interval, actual_threshold, emit_log
            )
            
            if stable_start:
                emit_log(f"找到满足条件的稳定时间段: {round(actual_threshold, 2)}K")
                break
            
            # 未找到，增加阈值继续尝试
            actual_threshold += temp_threshold_step
        
        # 记录实际使用的阈值
        if stable_start:
            emit_log(f"\n最终结果:")
            emit_log(f"  稳定时间间隔: {time_interval} 分钟")
            emit_log(f"  初始试验稳定温差: {round(initial_threshold, 2)}K")
            emit_log(f"  实际试验稳定温差: {round(actual_threshold, 2)}K")
        
        # 未能找到稳定条件
        if not stable_start:
            return {
                'message': f'未找到满足稳定条件的时间段\n稳定时间间隔: {time_interval}分钟\n初始试验稳定温差: {initial_threshold}K',
                'output_path': '',
                'warnings': self.warnings,
                'logs': self.logs
            }
        
        # 填写数据
        time_fmt = '%H:%M'
        
        # 填写 F 和 G 列
        for dr in data_rows:
            row_idx = dr['row']
            file_num, channel = dr['file'], dr['channel']
            
            temp_start = self.get_temp_at_time(file_num, channel, stable_start)
            temp_end = self.get_temp_at_time(file_num, channel, stable_end)
            
            red_fill = PatternFill(fill_type='solid', fgColor='FF0000', bgColor='FF0000')
            
            # F列 - 非合并单元格才写入和标记
            if not is_merged_cell(ws, row_idx, 6):
                if temp_start is not None:
                    ws.cell(row_idx, 6).value = round(temp_start, 1)
                else:
                    ws.cell(row_idx, 6).value = None
                    ws.cell(row_idx, 6).fill = red_fill
            
            # G列 - 非合并单元格才写入和标记
            if not is_merged_cell(ws, row_idx, 7):
                if temp_end is not None:
                    ws.cell(row_idx, 7).value = round(temp_end, 1)
                else:
                    ws.cell(row_idx, 7).value = None
                    ws.cell(row_idx, 7).fill = red_fill
        
        # 计算环境温度平均值
        f_vals = []
        g_vals = []
        
        for col_info in ambient_cols:
            row = col_info['row']
            f_val = ws.cell(row, 6).value
            g_val = ws.cell(row, 7).value
            
            # F列 - 非合并单元格才标记
            if f_val is not None:
                f_vals.append(f_val)
            elif not is_merged_cell(ws, row, 6):
                ws.cell(row, 6).fill = red_fill
            
            # G列 - 非合并单元格才标记
            if g_val is not None:
                g_vals.append(g_val)
            elif not is_merged_cell(ws, row, 7):
                ws.cell(row, 7).fill = red_fill
        
        f_avg = round(sum(f_vals) / len(f_vals), 1) if f_vals else 0
        g_avg = round(sum(g_vals) / len(g_vals), 1) if g_vals else 0
        
        # 填写平均值
        ws.cell(3, 6).value = f_avg  # F列
        ws.cell(3, 7).value = g_avg  # G列
        ws.cell(4, 6).value = stable_start.strftime(time_fmt)  # F列
        ws.cell(4, 7).value = stable_end.strftime(time_fmt)  # G列
        
        yellow_fill = PatternFill(fill_type='solid', fgColor='FFFF00', bgColor='FFFF00')
        red_fill = PatternFill(fill_type='solid', fgColor='FF0000', bgColor='FF0000')
        
        # 环境温度行的行号集合
        ambient_rows = {col_info['row'] for col_info in ambient_cols}
        
        # 计算温差并填写 H 列，同时检查相关条件
        for dr in data_rows:
            row_idx = dr['row']
            g_val = ws.cell(row_idx, 7).value
            limit_val = dr.get('limit')
            
            # 如果是环境温度行，H列填 "-"
            if row_idx in ambient_rows:
                ws.cell(row_idx, 8).value = "-"  # H列 - 环境温度行填 "-"
                continue
            
            if g_val is not None:
                delta = round(g_val - g_avg, 1)
                ws.cell(row_idx, 8).value = delta  # H列
                
                # 检查负数 - 填充黄色
                if delta < 0:
                    ws.cell(row_idx, 8).fill = yellow_fill
                
                # 检查 Rise > Limit（ref除外）
                b_value = dr.get('b_value', '')
                if limit_val is not None and delta > limit_val:
                    # 不是ref的行才标记红色
                    i_val = ws.cell(row_idx, 9).value  # I列
                    i_str = str(i_val).strip().lower() if i_val else ''
                    if 'ref' not in i_str and i_str not in ['--', '—', '-']:
                        ws.cell(row_idx, 8).fill = red_fill
                        self.warnings.append(f"行{row_idx}: Rise({delta}) > Limit({limit_val}) - {b_value}")
        
        # 保存文件
        output_dir = Path(excel_file).parent
        output_path = output_dir / "final-done.xlsx"
        
        wb.save(output_path)
        
        return {
            'message': f'处理完成！\n稳定时间: {stable_start.strftime("%H:%M")} - {stable_end.strftime("%H:%M")}\n稳定时间间隔: {time_interval}分钟\n初始试验稳定温差: {round(initial_threshold, 2)}K\n实际试验稳定温差: {round(actual_threshold, 2)}K\n环境温度均值: {f_avg}°C / {g_avg}°C\n输出文件: final-done.xlsx',
            'output_path': str(output_path),
            'warnings': self.warnings,
            'logs': self.logs
        }
    
    def parse_csv_directory(self, root_dir: str) -> list:
        """
        解析 CSV 目录结构
        
        目录结构：
        root_dir/
            1/
                xxx.csv  (只能有1个)
            2/
                yyy.csv
            3/
                zzz.csv
        
        Returns:
            list: ["/path/to/1/xxx.csv", "/path/to/2/yyy.csv", ...]
        
        Raises:
            ValueError: 目录结构不满足要求
        """
        root_path = Path(root_dir)
        
        if not root_path.is_dir():
            raise ValueError(f"目录检查阶段: 目录不存在 - {root_dir}")
        
        csv_files = []
        subdirs = sorted([d for d in root_path.iterdir() if d.is_dir()], key=lambda x: x.name)
        
        for subdir in subdirs:
            # 检查子目录名是否为数字
            if not subdir.name.isdigit():
                continue  # 跳过非数字命名的子目录
            
            # 查找该子目录下的 CSV 文件
            csv_list = list(subdir.glob("*.csv"))
            
            if len(csv_list) == 0:
                raise ValueError(f"目录检查阶段: 子目录 {subdir.name} 中没有 CSV 文件")
            elif len(csv_list) > 1:
                raise ValueError(f"目录检查阶段: 子目录 {subdir.name} 中有多个 CSV 文件，应只有1个")
            
            csv_files.append(str(csv_list[0]))
        
        if not csv_files:
            raise ValueError("目录检查阶段: 未找到任何 CSV 子目录（目录名应为数字，如1,2,3）")
        
        # 按子目录名排序返回
        csv_files.sort(key=lambda x: int(Path(x).parent.name))
        return csv_files
    
    def parse_csv(self, csv_path: str) -> tuple:
        """
        解析 CSV 文件
        
        Returns:
            tuple: (data_dict, min_time, max_time)
        """
        data = {}
        all_times = []
        
        # 尝试多种编码
        encodings = ['utf-16', 'utf-8', 'gbk', 'gb2312']
        
        for encoding in encodings:
            try:
                with open(csv_path, 'r', encoding=encoding) as f:
                    lines = f.readlines()
                break
            except:
                continue
        else:
            raise ValueError(f"CSV解析阶段: 无法读取 CSV 文件（编码错误）- {csv_path}")
        
        # 找到数据开始行
        data_start_line = 0
        for i, line in enumerate(lines):
            if '扫描' in line:
                data_start_line = i + 1
                break
        
        if data_start_line >= len(lines):
            raise ValueError(f"CSV解析阶段: CSV 文件格式不正确，找不到数据开始行 - {csv_path}")
        
        # 解析表头
        header_line = lines[data_start_line]
        headers = header_line.strip().split('\t')
        
        time_col = None
        channel_cols = {}
        
        for i, h in enumerate(headers):
            h = h.strip()
            if '时间' in h:
                time_col = i
            match = re.match(r'^(\d+)\s*\(C\)', h)
            if match:
                channel_cols[int(match.group(1))] = i
        
        if time_col is None:
            raise ValueError(f"CSV解析阶段: CSV 文件中没有找到时间列 - {csv_path}")
        
        if not channel_cols:
            raise ValueError(f"CSV解析阶段: CSV 文件中没有找到温度通道列 - {csv_path}")
        
        # 初始化数据
        for ch in channel_cols:
            data[ch] = []
        
        # 解析数据
        for line in lines[data_start_line + 1:]:
            if not line.strip():
                continue
            
            cols = line.strip().split('\t')
            if time_col is None or time_col >= len(cols):
                continue
            
            time_str = cols[time_col].strip()
            time_val = self.parse_time_string(time_str)
            
            if time_val is None:
                continue
            
            all_times.append(time_val)
            
            for ch, col_idx in channel_cols.items():
                if col_idx < len(cols):
                    try:
                        val = float(cols[col_idx].strip())
                        if val > -1e10 and val < 1e10:
                            data[ch].append((time_val, val))
                    except:
                        pass
        
        # 排序
        for ch in data:
            data[ch].sort(key=lambda x: x[0])
        
        min_time = min(all_times) if all_times else None
        max_time = max(all_times) if all_times else None
        
        return data, min_time, max_time
    
    def parse_time_string(self, time_str: str):
        """解析时间字符串"""
        # 格式: 2024/1/15 10:30:15:0
        match = re.match(r'(\d{4})/(\d+)/(\d+)\s+(\d+):(\d+):(\d+):(\d+)', time_str.strip())
        if match:
            year, month, day, hour, minute, second, ms = match.groups()
            return datetime(
                int(year), int(month), int(day),
                int(hour), int(minute), int(second),
                int(ms) * 1000
            )
        return None
    
    def get_temp_at_time(self, file_num: int, channel: int, target_time: datetime, tolerance: int = 60):
        """获取指定时间的温度值"""
        if file_num not in self.csv_data or channel not in self.csv_data[file_num]:
            return None
        
        data_points = self.csv_data[file_num][channel]
        if not data_points:
            return None
        
        times = [p[0] for p in data_points]
        
        # 找最近的时间点
        closest_idx = min(
            range(len(times)),
            key=lambda i: abs((times[i] - target_time).total_seconds())
        )
        
        if abs((times[closest_idx] - target_time).total_seconds()) <= tolerance:
            return data_points[closest_idx][1]
        
        return None
    
    def get_ambient_at_time(self, ambient_info: list, target_time: datetime):
        """获取指定时间的环境温度"""
        temps = []
        
        for file_num, channel in ambient_info:
            temp = self.get_temp_at_time(file_num, channel, target_time)
            if temp is not None:
                temps.append(temp)
        
        return sum(temps) / len(temps) if temps else None
    
    def find_stable_time(self, sorted_times: list, ambient_info: list, 
                         check_channels: list, time_interval: int, temp_threshold: float,
                         log_callback=None):
        """查找稳定时间段
        
        缓存优化：缓存温度查询结果避免重复计算
        """
        # 使用默认的日志函数（不实时输出）
        if log_callback is None:
            def emit_log(msg):
                self.logs.append(msg)
            log_callback = emit_log
        
        times_set = set(sorted_times)
        interval_seconds = time_interval * 60
        
        # 温度缓存：{(file_num, channel, time_key): temp_value}
        # time_key 使用秒数作为键
        temp_cache = {}
        
        # 辅助函数：从缓存获取温度
        def get_temp_cached(file_num, channel, target_time):
            time_key = int(target_time.timestamp())
            cache_key = (file_num, channel, time_key)
            
            if cache_key in temp_cache:
                return temp_cache[cache_key]
            
            temp = self.get_temp_at_time(file_num, channel, target_time)
            temp_cache[cache_key] = temp
            return temp
        
        # 辅助函数：从缓存获取环境温度
        def get_ambient_cached(target_time):
            time_key = int(target_time.timestamp())
            cache_key = ('ambient', time_key)
            
            if cache_key in temp_cache:
                return temp_cache[cache_key]
            
            temp = self.get_ambient_at_time(ambient_info, target_time)
            temp_cache[cache_key] = temp
            return temp
        
        log_callback(f"开始查找稳定时间（温差阈值: {round(temp_threshold, 2)}K，检查所有时间点）...")
        
        for t in sorted_times:
            t_interval_ago = t - timedelta(seconds=interval_seconds)
            
            # 使用set进行O(1)查找
            if t_interval_ago not in times_set:
                continue
            
            # 获取间隔前后的环境温度（使用缓存）
            ambient_before = get_ambient_cached(t_interval_ago)
            ambient_after = get_ambient_cached(t)
            
            if ambient_before is None or ambient_after is None:
                continue
            
            # 检查所有通道是否稳定
            is_stable = True
            current_max_diff = 0  # 当前时间点所有通道的最大温升差
            
            for file_num, channel in check_channels:
                temp_before = get_temp_cached(file_num, channel, t_interval_ago)
                temp_after = get_temp_cached(file_num, channel, t)
                
                if temp_before is None or temp_after is None:
                    continue
                
                # 计算相对温度变化
                diff = abs((temp_after - ambient_after) - (temp_before - ambient_before))
                
                # 记录最大差值
                if diff > current_max_diff:
                    current_max_diff = diff
                
                if diff >= temp_threshold:
                    is_stable = False
                    # 如果不满足阈值，记录这个通道的差异用于统计
                    break
            
            if is_stable:
                log_callback(f"✓ 找到稳定时间: {t_interval_ago.strftime('%H:%M')} - {t.strftime('%H:%M')} (温升差={round(current_max_diff, 2)}K)")
                return t_interval_ago, t
        
        log_callback(f"遍历完成，未找到满足条件 (<{round(temp_threshold, 2)}K) 的稳定时间段")
        
        return None, None
