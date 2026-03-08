"""
Tester Tool - Excel 验证模块
"""

from pathlib import Path
from openpyxl import load_workbook


class ExcelValidator:
    """Excel 文件验证器"""
    
    def __init__(self):
        self.errors = []
    
    def validate(self, excel_path: str, file_index_col: int = 4, channel_index_col: int = 5) -> dict:
        """
        验证 Excel 文件
        
        Args:
            excel_path: Excel 文件路径
            file_index_col: 文件索引列号（默认4=D列）
            channel_index_col: 通道索引列号（默认5=E列）
            
        Returns:
            dict: {
                'valid': bool,
                'message': str,
                'ambient_rows': list
            }
        """
        self.errors = []
        
        try:
            # 使用 data_only=False 以避免缓存问题
            wb = load_workbook(excel_path, data_only=False)
            ws = wb.active
            
            # 检查 F 和 G 列是否有重复
            f_col_values = []
            g_col_values = []
            
            # 确保工作表有足够的行
            if ws.max_row < 4:
                return {
                    'valid': False,
                    'message': '文件校验阶段: Excel 文件行数不足，至少需要4行数据',
                    'ambient_rows': []
                }
            
            for row in range(4, ws.max_row + 1):
                f_val = ws.cell(row, 6).value  # F列
                g_val = ws.cell(row, 7).value  # G列
                
                if f_val is not None and str(f_val).strip():
                    f_col_values.append((row, f_val))
                if g_val is not None and str(g_val).strip():
                    g_col_values.append((row, g_val))
            
            # 检查 F 列重复
            f_seen = {}
            for row, val in f_col_values:
                val_str = str(val).strip()
                if val_str in f_seen:
                    self.errors.append(f"F列重复: 行 {row} 和 {f_seen[val_str]} 都有值 {val_str}")
                else:
                    f_seen[val_str] = row
            
            # 检查 G 列重复
            g_seen = {}
            for row, val in g_col_values:
                val_str = str(val).strip()
                if val_str in g_seen:
                    self.errors.append(f"G列重复: 行 {row} 和 {g_seen[val_str]} 都有值 {val_str}")
                else:
                    g_seen[val_str] = row
            
            # 如果有错误，返回
            if self.errors:
                return {
                    'valid': False,
                    'message': '文件校验阶段: F列或G列存在重复值\n' + '\n'.join(self.errors),
                    'ambient_rows': []
                }
            
            # 查找所有数据行（列出B列所有名字，让用户选择哪些是环境温度）
            # 同时检查编号是否重复
            ambient_rows = []
            seen_file_channel = {}  # 用于检测 (file_num, channel) 重复
            
            for row in range(4, ws.max_row + 1):
                b_val = ws.cell(row, 2).value  # B列 - 名称
                d_val = ws.cell(row, file_index_col).value  # 用户选择的文件索引列
                e_val = ws.cell(row, channel_index_col).value  # 用户选择的通道索引列
                limit_val = ws.cell(row, 10).value  # J列 - Limit
                
                # 只要D列和E列有值就检查（包括B列为空的情况）
                if d_val is not None and e_val is not None:
                    try:
                        # 只有当 D 和 E 列都是有效数字时才处理
                        file_num = int(float(str(d_val)))
                        channel = int(float(str(e_val)))
                        
                        # 检查编号是否重复
                        key = (file_num, channel)
                        if key in seen_file_channel:
                            prev_row = seen_file_channel[key]
                            return {
                                'valid': False,
                                'message': f'文件校验阶段: 子目录索引{file_num}，通道号{channel}编号重复了（行{prev_row}和行{row}）',
                                'ambient_rows': []
                            }
                        seen_file_channel[key] = row
                        
                        # 只有B列也有值时才添加到ambient_rows（这是用户需要看到的行）
                        if b_val:
                            ambient_rows.append({
                                'row': row,
                                'b_value': str(b_val).strip(),  # B列名字
                                'd_value': d_val,
                                'e_value': e_val,
                                'limit': limit_val if limit_val is not None else ''  # J列 - Limit
                            })
                    except (ValueError, TypeError):
                        pass
            
            return {
                'valid': True,
                'message': '验证通过',
                'ambient_rows': ambient_rows  # 所有数据行，由用户选择哪些是环境温度
            }
            
        except Exception as e:
            return {
                'valid': False,
                'message': f'文件校验阶段: 验证失败 - {str(e)}',
                'ambient_rows': []
            }
    
    def get_errors(self) -> list:
        """获取错误列表"""
        return self.errors
